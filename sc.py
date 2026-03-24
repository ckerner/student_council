#!/usr/bin/env python3

import curses
import json
import os
import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

DATA_FILE = "sc_points.json"

# ----------------------------
# Data Layer
# ----------------------------

def load_data():
    if not os.path.exists(DATA_FILE):
        return {"students": [], "events": [], "attendance": []}
    with open(DATA_FILE, "r") as f:
        return json.load(f)


def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

# ----------------------------
# Helpers
# ----------------------------

def parse_date_input(date_str):
    date_str = date_str.strip()
    if date_str in (".", ""):
        return datetime.now().strftime("%Y-%m-%d")
    if date_str.startswith('-') and date_str[1:].isdigit():
        return (datetime.now() - timedelta(days=int(date_str[1:]))).strftime("%Y-%m-%d")
    if date_str.startswith('+') and date_str[1:].isdigit():
        return (datetime.now() + timedelta(days=int(date_str[1:]))).strftime("%Y-%m-%d")
    return date_str


def get_next_event_id(data):
    return max([e['id'] for e in data['events']], default=0) + 1


def find_event(data, event_id):
    return next((e for e in data['events'] if e['id'] == event_id), None)


def find_student(data, email):
    return next((s for s in data['students'] if s['email'] == email), None)

# ----------------------------
# Spreadsheet Export
# ----------------------------

def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2


def export_spreadsheet(data):
    wb = Workbook()

    totals = {s['email']: 0 for s in data['students']}
    for a in data['attendance']:
        e = find_event(data, a['event_id'])
        if e:
            totals[a['email']] += e['points']

    # Students
    ws1 = wb.active
    ws1.title = "Students"
    ws1.append(["Name", "Email", "Grade", "Points"])
    for cell in ws1[1]: cell.font = Font(bold=True)
    for s in data['students']:
        ws1.append([s['name'], s['email'], s['grade'], totals[s['email']]])
    autosize(ws1)
    ws1.freeze_panes = "A2"

    # Events
    ws2 = wb.create_sheet("Events")
    ws2.append(["ID", "Date", "Description", "Points"])
    for cell in ws2[1]: cell.font = Font(bold=True)
    for e in sorted(data['events'], key=lambda x: x['date']):
        ws2.append([e['id'], e['date'], e['description'], e['points']])
    autosize(ws2)
    ws2.freeze_panes = "A2"

    # Attendance
    ws3 = wb.create_sheet("Attendance")
    ws3.append(["Event ID", "Event", "Student", "Email"])
    for cell in ws3[1]: cell.font = Font(bold=True)
    for a in data['attendance']:
        e = find_event(data, a['event_id'])
        s = find_student(data, a['email'])
        if e and s:
            ws3.append([e['id'], e['description'], s['name'], s['email']])
    autosize(ws3)
    ws3.freeze_panes = "A2"

    # Leaderboard
    ws4 = wb.create_sheet("Leaderboard")
    ws4.append(["Rank", "Name", "Grade", "Points"])
    for cell in ws4[1]: cell.font = Font(bold=True)
    ranked = sorted(data['students'], key=lambda s: totals[s['email']], reverse=True)
    for i, s in enumerate(ranked, start=1):
        ws4.append([i, s['name'], s['grade'], totals[s['email']]])
    autosize(ws4)
    ws4.freeze_panes = "A2"

    # Summary
    ws5 = wb.create_sheet("Summary")
    ws5.append(["Metric", "Value"])
    for cell in ws5[1]: cell.font = Font(bold=True)
    ws5.append(["Total Students", len(data['students'])])
    ws5.append(["Total Events", len(data['events'])])
    ws5.append(["Total Attendance Records", len(data['attendance'])])
    autosize(ws5)

    wb.save("student_council_points.xlsx")

# ----------------------------
# CLI Reports
# ----------------------------

def report_students(data):
    totals = {s['email']: 0 for s in data['students']}
    for a in data['attendance']:
        e = find_event(data, a['event_id'])
        if e:
            totals[a['email']] += e['points']

    print(f"{'Name':30} {'Gr':>2} {'Email':25} {'Pts':>6}")
    print('-' * 70)
    for s in data['students']:
        print(f"{s['name']:30} {s['grade']:>2} {s['email']:25} {totals[s['email']]:>6}")


def report_events(data):
    print(f"{'ID':3} {'Date':12} {'Description':50} {'Pts':>6}")
    print('-' * 75)
    for e in data['events']:
        print(f"{e['id']:3} {e['date']:12} {e['description'][:50]:50} {e['points']:>6}")


def report_event_detail(data):
    for e in data['events']:
        print(f"Event {e['id']} - {e['description']} ({e['points']} pts, {e['date']})")
        attendees = [a['email'] for a in data['attendance'] if a['event_id'] == e['id']]
        if not attendees:
            print("   No attendees")
        for email in attendees:
            s = find_student(data, email)
            if s:
                print(f"   {s['name']:30} {s['grade']:>2} {s['email']}")


def report_student_detail(data):
    for s in data['students']:
        print(f"Student: {s['name']:30} {s['grade']:>2} {s['email']}")
        events = [a['event_id'] for a in data['attendance'] if a['email'] == s['email']]
        if not events:
            print("   No events attended")
        for eid in events:
            e = find_event(data, eid)
            if e:
                print(f"   Event {e['id']:3} {e['date']} {e['description'][:50]:50} {e['points']:>6}")
        print()

# ----------------------------
# TUI
# ----------------------------

def draw_list(stdscr, title, items, index):
    stdscr.clear()
    stdscr.addstr(0, 0, title)
    for i, item in enumerate(items):
        if i == index:
            stdscr.attron(curses.A_REVERSE)
        stdscr.addstr(i + 2, 0, item)
        if i == index:
            stdscr.attroff(curses.A_REVERSE)
    stdscr.refresh()


def edit_student(stdscr, s, data):
    curses.curs_set(1)
    curses.echo()
    stdscr.addstr(20, 0, f"Name [{s['name']}]: ")
    name = stdscr.getstr().decode()
    stdscr.addstr(21, 0, f"Email [{s['email']}]: ")
    email = stdscr.getstr().decode()
    stdscr.addstr(22, 0, f"Grade [{s['grade']}]: ")
    grade = stdscr.getstr().decode()
    curses.noecho()
    curses.curs_set(0)

    if name:
        s['name'] = name
    if grade:
        s['grade'] = int(grade)
    if email and email != s['email']:
        old_email = s['email']
        s['email'] = email
        for a in data['attendance']:
            if a['email'] == old_email:
                a['email'] = email

    save_data(data)


def edit_event(stdscr, e, data):
    curses.curs_set(1)
    curses.echo()
    stdscr.addstr(20, 0, f"Date [{e['date']}]: ")
    date_input = stdscr.getstr().decode()
    stdscr.addstr(21, 0, f"Description [{e['description']}]: ")
    desc = stdscr.getstr().decode()
    stdscr.addstr(22, 0, f"Points [{e['points']}]: ")
    pts = stdscr.getstr().decode()
    curses.noecho()
    curses.curs_set(0)

    if date_input:
        e['date'] = parse_date_input(date_input)
    if desc:
        e['description'] = desc
    if pts:
        e['points'] = int(pts)

    save_data(data)


def assign_event_to_students(stdscr, event, data):
    index = 0
    while True:
        items = [
            f"{'[X]' if any(a['email'] == s['email'] and a['event_id'] == event['id'] for a in data['attendance']) else '[ ]'} {s['name']:30} {s['grade']:>2} {s['email']:25}"
            for s in data['students']
        ]
        draw_list(stdscr, f"Assign Event {event['id']} - {event['description']}", items, index)

        key = stdscr.getch()
        if key in (ord('q'), 27):
            break
        elif key in (curses.KEY_DOWN, ord('j')):
            index = min(index + 1, len(items) - 1)
        elif key in (curses.KEY_UP, ord('k')):
            index = max(index - 1, 0)
        elif key in (10, 13):
            s = data['students'][index]
            existing = next((a for a in data['attendance'] if a['email'] == s['email'] and a['event_id'] == event['id']), None)
            if existing:
                data['attendance'].remove(existing)
            else:
                data['attendance'].append({'email': s['email'], 'event_id': event['id']})
            save_data(data)


def tui_main(stdscr, data):
    curses.curs_set(0)

    # Start in students view if none exist
    view = 'students' if not data['students'] else 'events'
    index = 0

    while True:
        if view == 'events':
            items = [f"{e['id']:3} {e['date']:12} {e['description'][:50]:50} {e['points']:>6}" for e in data['events']]
            if not items:
                items = ['No events']
            title = "Events (Tab=Switch View) - e=edit, a=add, d=delete, s=assign, q=quit"
        else:
            totals = {s['email']: 0 for s in data['students']}
            for a in data['attendance']:
                e = find_event(data, a['event_id'])
                if e:
                    totals[a['email']] += e['points']
            items = [f"{s['name']:30} {s['grade']:>2} {s['email']:25} {totals[s['email']]:>6}" for s in data['students']]
            if not items:
                items = ['No students']
            title = "Students (Tab=Switch View) - e=edit, a=add, d=delete, q=quit"

        draw_list(stdscr, title, items, index)

        key = stdscr.getch()

        if key in (ord('q'), 27):
            break
        elif key == 9:
            view = 'students' if view == 'events' else 'events'
            index = 0
        elif key in (curses.KEY_DOWN, ord('j')):
            index = min(index + 1, len(items) - 1)
        elif key in (curses.KEY_UP, ord('k')):
            index = max(index - 1, 0)
        elif key == ord('a'):
            if view == 'events':
                # Inline prompt for new event (no dummy row)
                curses.curs_set(1)
                curses.echo()
                stdscr.clear()
                stdscr.addstr(0,0,"Add Event (leave blank to cancel)")
                stdscr.addstr(2,0,"Date (. for today, +/-N): ")
                date_input = stdscr.getstr().decode().strip()
                stdscr.addstr(3,0,"Description: ")
                desc = stdscr.getstr().decode().strip()
                stdscr.addstr(4,0,"Points: ")
                pts = stdscr.getstr().decode().strip()
                curses.noecho()
                curses.curs_set(0)

                if desc:
                    try:
                        points = int(pts) if pts else 0
                    except ValueError:
                        points = 0
                    data['events'].append({
                        'id': get_next_event_id(data),
                        'date': parse_date_input(date_input or '.'),
                        'description': desc,
                        'points': points
                    })
                    save_data(data)
            else:
                # Inline prompt for new student (no dummy row)
                curses.curs_set(1)
                curses.echo()
                stdscr.clear()
                stdscr.addstr(0,0,"Add Student (leave blank name to cancel)")
                stdscr.addstr(2,0,"Name: ")
                name = stdscr.getstr().decode().strip()
                stdscr.addstr(3,0,"Email: ")
                email = stdscr.getstr().decode().strip()
                stdscr.addstr(4,0,"Grade (9-12): ")
                grade = stdscr.getstr().decode().strip()
                curses.noecho()
                curses.curs_set(0)

                if name:
                    try:
                        g = int(grade) if grade else 9
                        if g not in (9,10,11,12):
                            g = 9
                    except ValueError:
                        g = 9
                    data['students'].append({
                        'name': name,
                        'email': email,
                        'grade': g
                    })
                    save_data(data)
        elif key == ord('d'):
            if view == 'events' and data['events']:
                data['events'].pop(index)
            elif view == 'students' and data['students']:
                data['students'].pop(index)
            if index > 0:
                index -= 1
            save_data(data)
        elif key == ord('e'):
            if view == 'events' and data['events']:
                edit_event(stdscr, data['events'][index], data)
            elif view == 'students' and data['students']:
                edit_student(stdscr, data['students'][index], data)
        elif key == ord('s') and view == 'events' and data['events']:
            assign_event_to_students(stdscr, data['events'][index], data)

# ----------------------------
# Main
# ----------------------------

def main(stdscr):
    data = load_data()
    tui_main(stdscr, data)

# ----------------------------
# CLI
# ----------------------------

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Student Council Points Manager")
    sub = parser.add_subparsers(dest='command')

    sub.add_parser('tui', help='Launch interactive TUI')
    sub.add_parser('students', help='Show student summary with points')
    sub.add_parser('events', help='Show all event info')
    sub.add_parser('event_detail', help='Show which students attended each event')
    sub.add_parser('student_detail', help='Show events attended by each student')
    sub.add_parser('export', help='Export polished spreadsheet report')

    args = parser.parse_args()
    data = load_data()

    if args.command == 'tui' or args.command is None:
        curses.wrapper(main)
    elif args.command == 'students':
        report_students(data)
    elif args.command == 'events':
        report_events(data)
    elif args.command == 'event_detail':
        report_event_detail(data)
    elif args.command == 'student_detail':
        report_student_detail(data)
    elif args.command == 'export':
        export_spreadsheet(data)

